from openpyxl import load_workbook, Workbook
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import MergedCell
import os


def parser_merged_cell(sheet: Worksheet, row, col):
    """
    Check whether it is a merged cell and get the value of the corresponding row and column cell.
    If it is a merged cell, take the value of the cell in the upper left corner of the merged area as the value of the current cell; otherwise, directly return the value of the cell
    : param sheet: current sheet object
    : param row: the row of the cell to be obtained
    : param col: the column of the cell to be obtained
    :return:
    """
    cell = sheet.cell(row=row, column=col)
    if isinstance(cell, MergedCell):  # judge whether the cell is a merged cell
        for merged_range in sheet.merged_cell_ranges:  # loop to find the merge range to which the cell belongs
            if cell.coordinate in merged_range:
                # Gets the cell in the upper left corner of the merge range and returns it as the value of the cell
                cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
    return cell


def get_xlsx(name, src_name, dst_name, order, resultDir):
    wb = load_workbook(name)
    ws_data = wb.get_sheet_by_name(src_name)
    if dst_name in wb.sheetnames:
        ws_result = wb.get_sheet_by_name(dst_name)
    else:
        ws_result = wb.create_sheet(dst_name)
    row1 = ['测试案例编号', '计划测试日期', '前置条件', '测试概述', '操作步骤名称', '步骤描述', '预期结果', '状态']

    # 添加表头
    for index in range(len(row1)):
        _ = ws_result.cell(row=1, column=index + 1, value=row1[index])

    # 获取功能列表
    functions = []
    for row in ws_data.iter_rows(min_col=5, max_col=5):
        for cell in row:
            functions.append(cell.value)

    # 获取预期列表
    expected = []
    for row in ws_data.iter_rows(min_col=6, max_col=6):
        for cell in row:
            expected.append(cell.value)

    # 填写测试案例编号
    for rowIndex in range(len(functions)):
        if functions[rowIndex] is None:
            continue
        _ = ws_result.cell(row=rowIndex + 2, column=1, value='g{}00{}'.format(order, rowIndex))

    # 填写前置条件
    for rowIndex in range(len(functions)):
        if functions[rowIndex] is None:
            continue
        value = ''
        for column in ws_data.iter_cols(max_col=3, min_row=rowIndex + 1, max_row=rowIndex + 1):
            for cell in column:
                new_cell = parser_merged_cell(ws_data, cell.row, cell.column)
                if value == '':
                    value = new_cell.value
                    continue
                value = value + '->' + new_cell.value

        _ = ws_result.cell(row=rowIndex + 2, column=3, value=value)
        value = ''

    # 测试概述
    for rowIndex in range(len(functions)):
        if functions[rowIndex] is None:
            continue
        value = ''
        for column in ws_data.iter_cols(min_col=4, max_col=5, min_row=rowIndex + 1, max_row=rowIndex + 1):
            for cell in column:
                new_cell = parser_merged_cell(ws_data, cell.row, cell.column)
                if value == '':
                    value = '当' + new_cell.value + '时，'
                    continue
                value = value + '测试' + new_cell.value + '的情况'

        _ = ws_result.cell(row=rowIndex + 2, column=4, value=value)
        value = ''

    # 操作步骤名称
    for rowIndex in range(len(functions)):
        if functions[rowIndex] is None:
            continue
        _ = ws_result.cell(row=rowIndex + 2, column=5, value=functions[rowIndex])

    # 步骤描述
    for rowIndex in range(len(functions)):
        if functions[rowIndex] is None:
            continue
        value = ''
        for column in ws_data.iter_cols(min_col=4, max_col=5, min_row=rowIndex + 1, max_row=rowIndex + 1):
            for cell in column:
                new_cell = parser_merged_cell(ws_data, cell.row, cell.column)
                if value == '':
                    value = new_cell.value
                    continue
                value = value + '->' + new_cell.value

        _ = ws_result.cell(row=rowIndex + 2, column=6, value=value)
        value = ''

    # 预期结果
    for rowIndex in range(len(expected)):
        if expected[rowIndex] is None:
            continue
        _ = ws_result.cell(row=rowIndex + 2, column=7, value=expected[rowIndex])

    wb.save(name)
    if not os.path.exists(resultDir):
        wbn = Workbook()
        wbn.save(resultDir)

    wb2 = load_workbook(resultDir)

    if dst_name in wb2.sheetnames:
        ws_result_new = wb2.get_sheet_by_name(dst_name)
    else:
        ws_result_new = wb2.create_sheet(dst_name)
    for value in ws_result.iter_rows(min_row=1, max_row=ws_result.max_row, min_col=1, max_col=ws_result.max_column, values_only=True):
        value = list(value)
        ws_result_new.append(value)

    if 'Sheet' in wb2.sheetnames:
        del wb2['Sheet']

    wb2.save(resultDir)




def reset_col(filename):
    wb = load_workbook(filename)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        df = pd.read_excel(filename, sheet).fillna('-')
        df.loc[len(df)] = list(df.columns)  # 把标题行附加到最后一行
        for col in df.columns:
            index = list(df.columns).index(col)  # 列序号
            letter = get_column_letter(index + 1)  # 列字母
            collen = df[col].apply(lambda x: len(str(x).encode())).max()  # 获取这一列长度的最大值 当然也可以用min获取最小值 mean获取平均值
            ws.column_dimensions[letter].width = collen * 1   # 也就是列宽为最大长度*1.2 可以自己调整

    wb.save(filename)




# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    baseDir = r"/Users/tianzc/同步文件夹/双向同步/honor14/论文/my/测试用例/"
    fList = os.listdir(baseDir)
    fileNameList = [file for file in fList if 'xlsx' in file]
    resultDir = './result.xlsx'
    for i in range(len(fileNameList)):
        get_xlsx(baseDir + fileNameList[i], fileNameList[i].replace('.xlsx', ''), fileNameList[i].replace('.xlsx', '测试用例'), i,resultDir )

    reset_col(resultDir)

